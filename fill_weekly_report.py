
import os
import shutil
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook #imports python library for reading and writting excel files
from openpyxl.drawing.image import Image
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict

# event_financial_reports_folder = Path("event_financial_reports")
# year = "2026"
# month = "02_February"
# weekly_report = "02-23-2026_weekly_report"
# week_folder_path = event_financial_reports_folder / year / month / weekly_report

# Extract info from weekly_data.json; Needs folder path provided by reate_weekly_report_folder function
def extract_weekly_data(week_folder_path, week_start):

    #weekly_data_json ="02-23-26_weekly_data.json"
    weekly_data_json =f"{week_start}_weekly_data.json"
    weekly_data_json_path = week_folder_path/ weekly_data_json
    #print(weekly_data_json_path.exists())
    with open(weekly_data_json_path, 'r') as file:
         #data is a list of dict
         event_info = json.load(file)

    weekly_data = event_info["weekly_data"]

    event_date= [data["event_date"] for data in weekly_data]
    event_name= [data["event_name"]for data in weekly_data]
    event_type= [data["event_type"]for data in weekly_data]
    guest_count= [int(data["guest_count"])for data in weekly_data]

    space_fee= [float(data["space_fee"])for data in weekly_data]
    food_revenue=[float(data["food_revenue"])for data in weekly_data]
    beverage_revenue= [float(data["beverage_revenue"])for data in weekly_data]
    admin_fee= [float(data["admin_fee"])for data in weekly_data]
    sales_tax= [float(data["sales_tax"])for data in weekly_data]
    gross_revenue=[float(data["gross_revenue"])for data in weekly_data]
    net_revenue=[float(data["gross_revenue"]) - float(data["sales_tax"])for data in weekly_data]

    additional_labor_cost = [float(data["additional_labor"])for data in weekly_data]

    in_house_labor_cost= [float(data["in_house_labor"])for data in weekly_data]
    
    outsourced_labor_cost= [float(data["outsourced_labor"])for data in weekly_data]

    total_labor_cost =[
                        float(data["in_house_labor"]) +float(data["additional_labor"]) + float(data["outsourced_labor"])
                        for data in weekly_data
    ]
    
    approx_food_cost= [float(data["approx_food_cost"])for data in weekly_data]
    approx_bev_cost =[(float(data["beverage_revenue"]) * 0.22 )for data in weekly_data]

    rentals= [float(data["rentals"])for data in weekly_data]

    approx_operating_cost =[float(afc) + float(abc) + float(tlc) + float(r) for afc, abc, tlc, r in zip(approx_food_cost, approx_bev_cost, total_labor_cost, rentals)
                             ]
    
    operating_profit = [float(nr) - float(aoe) for nr, aoe in zip(net_revenue, approx_operating_cost)]
    margin_pct =[float(op) / float(nr) for op, nr in zip(operating_profit, net_revenue)]
    revenue_per_guest =[float(nr) / float(gc) for nr, gc in zip(net_revenue, guest_count)]
    food_and_beverage_cost_per_guest =[(float(afc) + float(abc)) / float(gc) for afc, abc, gc in zip(approx_food_cost, approx_bev_cost, guest_count)]
    labor_cost_per_guest =[[float(tlc) / float(gc) for tlc, gc in zip(total_labor_cost, guest_count)]]
    profit_per_guest =[float(op) / float(gc) for op, gc in zip(operating_profit, guest_count)]

     #---------------------------------------------------------------------------------------------------------------

# Fills out xlsx weekly report file.

    col_count = len(weekly_data)
    row_start =1
    row_end = 29
    formula_col = 2  # Column B is the formula source column
    start_col = 3  # Column C is the first column to be copied to
    end_col = start_col + col_count - 2

    

    weekly_report_xlsx =f"{week_start}_weekly_report.xlsx"
    weekly_data_xlsx_path = week_folder_path/ weekly_report_xlsx

    wb =load_workbook(weekly_data_xlsx_path)
    ws_1 =wb['Event_PnL']

    # Iterate over cells and check for formula. If it exists, copy over to the adjacent cell in the next col.
    for row in range(row_start, row_end + 1):
        src_cell = ws_1.cell(row=row, column=formula_col)

        if src_cell.data_type == "f":
            for target_col in range(start_col, end_col + 1):
                dst_cell = ws_1.cell(row=row, column=target_col)
                dst_cell.value = Translator(
                    src_cell.value,
                    origin=src_cell.coordinate
                ).translate_formula(dst_cell.coordinate)
    
    full_data =[{
        "event_date":e_date, 
        "event_name":e_name, 
        "event_type":e_type, 
        "guest_count":gc, 
        "space_fee":sf, 
        "food_revenue":fr, 
        "beverage_revenue":br,
         "admin_fee": af,
         "sales_tax":st,
         "gross_revenue":gr,
         "additional_labor_cost":alc, 
         "in_house_hourly_labor_cost":ihl,
         "outsourced_labor_cost": ol, 
         "approx_food_cost":afc, 
         "rentals": r} for e_date, e_name, e_type, gc, sf,fr,br,af,st,gr,alc, ihl,ol,afc, r in zip(
        event_date, event_name, event_type, guest_count, space_fee, food_revenue, beverage_revenue, admin_fee, sales_tax, 
        gross_revenue, additional_labor_cost,in_house_labor_cost, outsourced_labor_cost,approx_food_cost, rentals
    )]
# Fills out the Event_PnL portion of weekly_report.xlsx 
    info_col_start =2 # Col B in Even_PnL sheet
    header_col = 1 # Col A in Even_PnL sheet
    for event_info in full_data:
        for row in range(row_start, row_end + 1):
            header_cell = ws_1.cell(row=row, column=header_col).value
            target_cell = ws_1.cell(row = row, column=info_col_start)

            if header_cell in event_info and target_cell.value in (None, ""):
                ws_1.cell(row=row, column=info_col_start).value = event_info[header_cell]

        info_col_start += 1
        

    # Fill out the invoices portion of weekly_report.xlsx 
    weekly_invoices_json =f"{week_start}_weekly_invoice.json"
    weekly_invoices_json_path = week_folder_path/ weekly_invoices_json

    
    ws_2 =wb['Invoices']

    with open(weekly_invoices_json_path, 'r') as file:
         #data is a list of dict
         invoices = json.load(file)
    weekly_invoices =invoices["weekly_invoices"]

    invoice_date =[data["invoice_date"] for data in weekly_invoices]
    vendor = [data["vendor"] for data in weekly_invoices]
    invoice_number = [data["invoice_number"] for data in weekly_invoices]
    category = [data["category"] for data in weekly_invoices]
    cost = [float(data["cost"])for data in weekly_invoices]
    notes =[data["notes"] for data in weekly_invoices]

    full_invoices =[
        {"invoice_date":in_d, "vendor":v, "invoice_number":in_n, "category":cat, "cost":cos, "notes":no}
        for in_d, v,in_n, cat,cos,no in zip(invoice_date, vendor, invoice_number, category, cost, notes)
    ]

    invoice_start_col = 1 # Col A in Event_PnL sheet
    invoice_end_col = 6 # Col F in Event_PnL sheet
    invoice_start_row = 2

    # for invoice_info in full_invoices:
    #     for col in range(invoice_start_col, invoice_end_col + 1):
    #         header_cell = ws_2.cell(row=1, column= invoice_start_col).value
    #         ws_2.cell(row=invoice_start_row, column=invoice_start_col).value = invoice_info[header_cell]
    #     invoice_start_col += 1

    for invoice_info in full_invoices:
        for col in range(invoice_start_col, invoice_end_col + 1):
            header_cell = ws_2.cell(row=1, column= col).value
            
            ws_2.cell(row=invoice_start_row, column=col).value = invoice_info[header_cell]
        invoice_start_row += 1

    wb.save(weekly_data_xlsx_path)

    #---------------------------------------------------------------------------------------------------------------
    # Charts folder
    chart_file_names = []

    charts_folder ="charts"
    charts_path = Path(week_folder_path / charts_folder)

    x_labels = [
    f"{name}\n{guests} guests\n{etype}"
    for name, guests, etype in zip(event_name, guest_count, event_type)
    ]
    # Chart Revenue vs Operating Profit (by event); Grouped Bar Chart

    x = np.arange(len(weekly_data))
    width = 0.4

    plt.figure()

    plt.bar(x - width/2, net_revenue, width, label="Net Revenue")
    plt.bar(x + width/2, operating_profit, width, label="Operating Profit")


    plt.xticks(x, x_labels, rotation=30)

    plt.ylabel("USD")
    plt.title("Revenue vs Operating Profit by Event")

    plt.legend()
    plt.tight_layout()

    plt.savefig(charts_path/ f"{week_start}_revenue_vs_profit.png", dpi=300, bbox_inches="tight")
    plt.close()
    chart_file_names.append(f"{week_start}_revenue_vs_profit.png")

    # Chart Revenue per Guest; Bar Chart

    plt.figure()

    plt.bar(event_name, revenue_per_guest)

    plt.ylabel("Revenue per Guest ($)")
    plt.title("Revenue per Guest by Event")

    plt.xticks(x, x_labels, rotation=30)
    plt.tight_layout()

    plt.savefig(charts_path / f"{week_start}_revenue_per_guest.png", dpi=300, bbox_inches="tight")
    plt.close()

    chart_file_names.append(f"{week_start}_revenue_per_guest.png")

    #---------------------------------------------------------------------------------------------------------------

    # Chart Operating Profit by Event Type; Bar Chart

    profit_by_type = defaultdict(float)

    for etype, profit in zip(event_type, operating_profit):
        profit_by_type[etype] += profit

    types = list(profit_by_type.keys())
    profits = list(profit_by_type.values())

    plt.figure()

    plt.bar(types, profits)

    plt.ylabel("Operating Profit ($)")
    plt.title("Operating Profit by Event Type")

    plt.xticks(x, x_labels, rotation=30)
    plt.tight_layout()

    plt.savefig(charts_path / f"{week_start}_profit_by_event_type.png", dpi=300, bbox_inches="tight")
    plt.close()

    chart_file_names.append(f"{week_start}_profit_by_event_type.png")

    # Chart Operating Cost by Event Type; Bar Chart

    cost_by_type = defaultdict(float)

    for etype, cost in zip(event_type, approx_operating_cost):
       cost_by_type[etype] += cost

    types = list(cost_by_type.keys())
    cost = list(cost_by_type.values())

    plt.figure()

    plt.bar(types, cost)

    plt.ylabel("Operating Cost ($)")
    plt.title("Operating Cost by Event Type")

    plt.xticks(x, x_labels, rotation=30)
    plt.tight_layout()

    plt.savefig(charts_path / f"{week_start}_cost_by_event_type.png", dpi=300, bbox_inches="tight")
    plt.close()

    chart_file_names.append(f"{week_start}_cost_by_event_type.png")

    # Chart Event Type vs Food + Bev + Labor Cost; Stacked Bar Chart

    food_by_type = defaultdict(float)
    bev_by_type = defaultdict(float)
    labor_by_type = defaultdict(float)
    rental_by_type = defaultdict(float)

    for etype, afc, abc, tlc, r in zip(event_type, approx_food_cost, approx_bev_cost, total_labor_cost, rentals):
        food_by_type[etype] += afc
        bev_by_type[etype] += abc
        labor_by_type[etype] += tlc
        rental_by_type[etype] += r



    types = list(food_by_type.keys())

    food_values = [food_by_type[t] for t in types]
    bev_values = [bev_by_type[t] for t in types]
    labor_values = [labor_by_type[t] for t in types]
    rental_values = [rental_by_type[t] for t in types]


    plt.figure()
    plt.bar(types, food_values, label="Food Cost")

    plt.bar(types, bev_values,
            bottom=food_values,
            label="Beverage Cost")

    bottom_stack = [
        f + b for f, b in zip(food_values, bev_values)
    ]

    plt.bar(types, labor_values,
            bottom=bottom_stack,
            label="Labor Cost")
    
    full_bottom_stack = [
        f + b +l for f, b, l in zip(food_values, bev_values, labor_values)
    ]

    plt.bar(types, rental_values, bottom=full_bottom_stack, label="Rentals")

   
    
    plt.ylabel("Cost ($)")
    plt.title("Food + Bev + Labor Cost by Event Type")

    plt.legend()
    plt.tight_layout()

    plt.xticks(x, x_labels, rotation=30)


    plt.savefig(charts_path / f"{week_start}_cost_breakdown_by_event_type.png", dpi=300,bbox_inches="tight")
    plt.close()

    chart_file_names.append(f"{week_start}_cost_breakdown_by_event_type.png")

    # Chart Weekly Operating Profit vs Weekly Operating Cost
    
    weekly_operating_cost = sum(approx_operating_cost)
    weekly_operating_profit = sum(operating_profit)

    labels = ["Operating Cost", "Operating Profit"]
    values = [weekly_operating_cost, weekly_operating_profit]

    plt.figure()

    plt.bar(labels, values)

    plt.ylabel("USD")
    plt.title("Weekly Operating Profit vs Cost")

    plt.tight_layout()

    plt.savefig(charts_path / f"{week_start}_weekly_profit_vs_cost.png", dpi=300, bbox_inches="tight")
    plt.close()

    chart_file_names.append(f"{week_start}_weekly_profit_vs_cost.png")
        
    print("✅ Charts Created!")

#--------------------------------------------------------------------------------------------------------------
    #Fills out weekly_report.xlsx with charts
    ws_3 = wb["Charts"]
    chart_start_row = 2
    row_jump = 23
    chart_start_col = 1
    max_charts_per_col = 2
    chart_count = 0
    for chart_name in chart_file_names:
        charts_folder_path = charts_path/ chart_name
        
        chart_start_col_letter = get_column_letter(chart_start_col)
        img = Image(charts_folder_path)
        img.width = 900
        img.height = 900
        ws_3.add_image(img, f"{chart_start_col_letter}{chart_start_row}")
        chart_count += 1        
        if chart_count < max_charts_per_col:
            chart_start_row += row_jump
            chart_start_col += 1
        else:
            chart_start_row =2
            chart_start_col += 3
            chart_count = 0

    wb.save(weekly_data_xlsx_path)

    print(f"✅ {week_start}_weekly_report.xlsx filled!")


#extract_weekly_data(week_folder_path,"02-23-2026")