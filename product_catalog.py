import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers, Font, PatternFill, Border, Side, Alignment
import sqlite3
import json
from fuzzy import match_menu_items, normalize, fuzzy_match
from datetime import date
from excel_format import format_table
from database import create_df
from openpyxl.utils import get_column_letter
from excel_format import format_table_ver_2



"""Updates ingredient table in chosen database"""
def update_ingredient_table(db_file, product_catalog_excel_file):

    source_file = product_catalog_excel_file
    df_source = pd.read_excel(source_file)  

    
    source_data = df_source[["Vendor Name","Item Code","Item Description","Normalized Item Description","Pack/Size/Unit","Last Purchased Price ($)","Product(s)"]]
    source_data_set = []
    
    # bad_rows = source_data[pd.isna(source_data["Vendor Name"])]
    # print(bad_rows)


    conn = sqlite3.connect(db_file)
    # Cursor to execute commands
    cursor = conn.cursor()
    cursor.execute("""CREATE UNIQUE INDEX IF NOT EXISTS idx_ingredients_vendor_code_pack
                    ON ingredients(purveyor, ingredient_code, pack_size_unit);""")
    
    conn.commit()

    for _, row in source_data.iterrows():
        #purveyor = (row.get("Vendor Name", "") or "").strip()
        purveyor = "" if pd.isna(row.get("Vendor Name")) else str(row.get("Vendor Name")).strip()

        raw_code = row.get("Item Code")

        if pd.isna(raw_code):
            code = ""
        else:
            code = str(raw_code).strip()
            if code.endswith(".0") and code[:-2].isdigit():
                code = code[:-2]

        cursor.execute("""
            INSERT INTO ingredients
                (purveyor, ingredient_code, ingredient_description, ingredient_name, pack_size_unit, purchase_price, ingredient_type)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(purveyor, ingredient_code, pack_size_unit) DO UPDATE SET
                ingredient_description = excluded.ingredient_description,
                ingredient_name        = excluded.ingredient_name,
                pack_size_unit         = excluded.pack_size_unit,
                purchase_price         = excluded.purchase_price,
                ingredient_type        = excluded.ingredient_type;
        """, (
            purveyor,
            code,
            row.get("Item Description",""),
            row.get("Normalized Item Description",""),
            row.get("Pack/Size/Unit",""),
            row.get("Last Purchased Price ($)", ""),
            row.get("Product(s)", "")
        ))


    conn.commit()
    conn.close()
    
    #print(data_to_upload)

    print(f"✅ Ingredients Table has been updated!")
    
#----------------------------------------------------------------------------------------
def input_menu_ingredient(db):

    """input new ingredients into ingredient table for desired database.  """

    file_path = "input_ingredient.json"
    conn = sqlite3.connect(db)
    # Cursor to execute commands
    cursor = conn.cursor()

    cursor.execute("""
                SELECT ingredient_id, purveyor, ingredient_code, ingredient_description, ingredient_name, pack_size_unit, purchase_price, ingredient_type
                FROM ingredients;
                  """)
    
    results = cursor.fetchall()

    db_ingredients=[{"ingredient_id":result[0],
                    "purveyor": result[1],  
                     "ingredient_code": result[2], 
                     "ingredient_description": result[3],
                     "ingredient_name":result[4] ,
                     "pack_size_unit":result[5], 
                     "purchase_price": result[6], 
                     "ingredient_type":result[7]} for result in results]
    #print(ingredient_names)

    if os.path.exists(file_path):
        print("file_path is correct")
    else:
        print("ERROR")
          
    # Read the existing content
    with open(file_path, 'r') as file:
        #data is a list of dict
        data = json.load(file)

    read_ing_file = data["ingredient"]
   
    """ Check ingredients for existing ingredient_name """

    ingredient_lookup = {
        (normalize(db_ing["purveyor"]),db_ing["ingredient_code"]): db_ing
        for db_ing in db_ingredients
        if db_ing["ingredient_code"]
        }

    ingredient_to_upload = []
    for ing in read_ing_file:
           
        ingredient_to_upload.append({"purveyor": normalize(ing["purveyor"]),  
                     "ingredient_code": ing["ingredient_code"], 
                     "ingredient_description": ing["ingredient_description"],
                     "ingredient_name":normalize(ing["ingredient_description"]) ,
                     "pack_size_unit":ing["pack_size_unit"], 
                     "purchase_price": ing["purchase_price"], 
                     "ingredient_type":ing["ingredient_type"]})

    for ing in ingredient_to_upload:    

        key = (normalize(ing["purveyor"]), ing["ingredient_code"])

        if key in ingredient_lookup:
            match = ingredient_lookup[key]

            if match:
                print(f"🍽️ ingredient exists: {ingredient_to_upload}")
                update_prompt = input(f"Would you like to overwrite {ingredient_to_upload} in {db}?: y/n  ")
                if update_prompt != "y":
                    continue
                update_sql = """
                                UPDATE ingredients
                                SET
                                    ingredient_description = :desc,
                                    ingredient_name        = :name,
                                    pack_size_unit         = :pack,
                                    purchase_price         = :price,
                                    ingredient_type        = :type
                                WHERE ingredient_id = :id;
                                """

                cursor.execute(update_sql, {
                                    "desc": ing["ingredient_description"],
                                    "name": ing["ingredient_name"],
                                    "pack": ing["pack_size_unit"],
                                    "price": ing["purchase_price"],
                                    "type": ing["ingredient_type"],
                                    "id": match["ingredient_id"],
                                })
                conn.commit()  
                print(f"✅ ingredient: {ing} overwritten!")

        
        else:
            print(f"{ingredient_to_upload} does not exist in: {db}")
            proceed_prompt = input(f"would you like to proceed?: y/n  ")
            if proceed_prompt != "y":
                continue
            else:
                cursor.execute("""INSERT OR IGNORE INTO ingredients 
                               (purveyor,
                                ingredient_code,
                                ingredient_description,
                                ingredient_name,
                                pack_size_unit, 
                                purchase_price, 
                                ingredient_type) VALUES (?,?,?,?,?,?,?)""", 
                               (ing["purveyor"],
                                ing["ingredient_code"],
                                ing["ingredient_description"],
                                ing["ingredient_name"],
                                ing["pack_size_unit"],
                                ing["purchase_price"],
                                ing["ingredient_type"],))
 
                conn.commit()
    conn.close()
   
    print(f"✅ ingredient: {ing} uploaded!")


#----------------------------------------------------------------------------------------
def get_menu_item_ingredients(db):

    conn = sqlite3.connect(db)
    # Cursor to execute commands
    cursor = conn.cursor()

    results= []
    final_results =[]
    cursor.execute("""
                SELECT menu_item_id, item_name 
                FROM menu_items;
                  """)
    
    menu_items = cursor.fetchall()
    for tuple_item in menu_items:
            results.append({"menu_item_id": tuple_item[0], "item_name": tuple_item[1], "ingredient_ids": [],
               "ingredient_name": "", "ingredient_code": "", "purveyor": "", "purchase_price": 0.0})

    for dict_item in results:
        cursor.execute("""
                SELECT ingredient_id
                FROM menu_ingredients
                WHERE menu_item_id = ?;
                """, (dict_item["menu_item_id"],)
            )
        
        
        ingredients = cursor.fetchall()
        for ingredient in ingredients:
            dict_item["ingredient_ids"].append(ingredient[0])
    
    for result in results:
        for id in result["ingredient_ids"]:
               final_results.append({
                "menu_item_id":result["menu_item_id"],
                "item_name":result["item_name"],
                "ingredient_id":id,
                })
               
    for result in final_results:
        cursor.execute("""
        SELECT ingredient_name, ingredient_code, purveyor, purchase_price
        FROM ingredients
        WHERE ingredient_id = ?;
    """, (result["ingredient_id"],))
    
        ingredient_data = cursor.fetchone()
   
        if ingredient_data:
            result["ingredient_name"] = ingredient_data[0]
            result["ingredient_code"] = ingredient_data[1]
            result["purveyor"] = ingredient_data[2]
            result["purchase_price"] = ingredient_data[3]

    # for result in final_results:     
    #     print(f"\n {result}")
    

    excel_file_count = 0
    # Create an excel file
    excel_file = f"updated_menu_ingredient_list.xlsx"

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file)
    ingredient_sheet= workbook["ingredients"]

    # # Iterate over each row and column in the sheet
    # Write each menu item + ingredient combination to the Excel sheet
    for row, dict_items in enumerate(final_results, start=2):
        ingredient_sheet.cell(row=row, column=1, value=dict_items.get("menu_item_id", ""))
        ingredient_sheet.cell(row=row, column=2, value=dict_items.get("item_name", ""))
        ingredient_sheet.cell(row=row, column=3, value=dict_items.get("ingredient_id", ""))
        ingredient_sheet.cell(row=row, column=4, value=dict_items.get("ingredient_name", ""))
        ingredient_sheet.cell(row=row, column=5, value=dict_items.get("ingredient_code", ""))
        ingredient_sheet.cell(row=row, column=6, value=dict_items.get("purveyor", ""))
        ingredient_sheet.cell(row=row, column=7, value=dict_items.get("purchase_price", 0.0))

    workbook.save(excel_file)
    print("✅ updated_menu_ingredient_list.xlsx has been updated!")

#----------------------------------------------------------------------------------------

def menu_cost(db):
    current_date = date.today()
    conn = sqlite3.connect(db)
    # Cursor to execute commands
    cursor = conn.cursor()
    current_date = date.today()
    menu_item_cost = []
    cursor.execute(
        """ SELECT *
            FROM menu_items;
        """)
    item_id_tuple = cursor.fetchall()
    for id in item_id_tuple:
       menu_item_cost.append({"item_id":id[0], "item_name":id[1], "ingredients":[]})

    #print(menu_item_cost)

    for item in menu_item_cost:
        cursor.execute(f"""
                       SELECT ingredients.ingredient_id,  ingredients.ingredient_name, ingredients.purveyor, ingredients.ingredient_code, ingredients.pack_size_unit, 
                       CAST (ingredients.purchase_price AS FLOAT)
                       FROM ingredients
                       JOIN menu_ingredients ON ingredients.ingredient_id = menu_ingredients.ingredient_id
                       WHERE menu_ingredients.menu_item_id = {item["item_id"]};
                       """)  
        ingredient_tuple = cursor.fetchall()
        #print(ingredient_tuple)
        for ing_tuple in ingredient_tuple:
            try:
                item["ingredients"].append({"ingredient_id":ing_tuple[0],
                                        "ingredient_name":ing_tuple[1],
                                        "purveyor": ing_tuple[2],
                                        "ingredient_code": ing_tuple[3],
                                        "pack_size_unit": ing_tuple[4],
                                        "price": ing_tuple[5]})
            
            except:
                continue
    conn.close()

    for item in menu_item_cost:
        item["ingredients"] = create_df(item["ingredients"])
        
    print(menu_item_cost)

 
    # Create an excel file
    excel_file_count = 0
    excel_file = f"Events_Menu_Cost_{excel_file_count}_{current_date}.xlsx"

   
    
    # Continously checks until it finds a non-existent file name
    while os.path.exists(excel_file):
        excel_file_count += 1
        # This updates the file_count, allowing for it to be checked again in the while loop
        excel_file = f"Events_Menu_Cost_{excel_file_count}_{current_date}.xlsx"
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "menu_cost"
        wb.save(excel_file)
        print(f"✅ {excel_file} Created!")

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
        item_name_row = 3
        item_id_row = 4
        df_row = 4
        current_col = 0  # 0-based for pandas
        column_A = 1 #1-based for openpyxl (col A)
        column_F = 6 #6-based for openpyxl (col F)
        column_E = 5
        pd.DataFrame().to_excel(writer, sheet_name='menu_cost', index=False)
        for item_dict in menu_item_cost:

            current_col_letter = get_column_letter(column_A)
            end_col_letter = get_column_letter(column_F)

            ws = writer.sheets['menu_cost']

            # prepares cells for item_name
            cell_item_name = ws[f"{current_col_letter}{item_name_row}"]
            cell_item_name.value = item_dict['item_name'].capitalize()
            cell_item_name.border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
            cell_item_name.alignment = Alignment(horizontal='center', vertical='center')
            cell_item_name.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
            cell_item_name.font = Font(name='Calibri', size=14,bold=True, color="FF000000")
            ws.merge_cells(f"{current_col_letter}{item_name_row}:{end_col_letter}{item_name_row}")

            # prepares cells for item_id
            cell_item_id =  ws[f"{current_col_letter}{item_id_row}"]
            cell_item_id.value = f'Item_ID: {item_dict["item_id"]}'
            cell_item_id.border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
            cell_item_id.alignment = Alignment(horizontal='center', vertical='center')
            cell_item_id.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
            cell_item_id.font = Font(name='Calibri', size=14,bold=True, color="FF000000")
            ws.merge_cells(f"{current_col_letter}{item_id_row}:{end_col_letter}{item_id_row}")


            # populates cells with ingredients
            item_dict['ingredients'].to_excel(writer, sheet_name= 'menu_cost', startrow=df_row, startcol=current_col, index=False)

            format_table_ver_2(ws, df_row, column_A, item_dict['ingredients'])

            df_length = len(item_dict['ingredients'])
            first_price_cell =f"{end_col_letter}{df_row + 2}"
            last_price_cell = f"{end_col_letter}{df_row + df_length + 1}"

            
            sum_row = df_row + len(item_dict['ingredients'])  + 2
            # print(f"df_length: {df_length}")
            #print(f"Sum Row: {sum_row}")
            # sum_cell = ws[f"{end_col_letter}{sum_row}"]
            total_sum_cost = ws.cell(row=sum_row, column=column_F)
            total_sum_cost.value = f"=SUM({first_price_cell}:{last_price_cell})"
            total_sum_cost.border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
            total_sum_cost.alignment = Alignment(horizontal='center', vertical='center')
            total_sum_cost.font = Font(bold=True, name='Calibri', size=12, color="000000")


            total_cost_string = ws.cell(row=sum_row, column=column_E)
            total_cost_string.value = "TOTAL COST"
            total_cost_string.border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
            total_cost_string.alignment = Alignment(horizontal='center', vertical='center')
            total_cost_string.font = Font(bold=True, name='Calibri', size=12, color="000000")

            #format_table(ws, df_row, current_category_col, item_dict['ingredients'])
            #this is openpyxl based and starts at 1 for indexing
            df_row += len(item_dict["ingredients"]) + 5 # Add space between tables
            item_id_row += len(item_dict["ingredients"]) + 5
            item_name_row += len(item_dict["ingredients"]) + 5
            #insert_blank_rows(ws, df_row + 1 )
          
#----------------------------------------------------------------------------------------

def menu_cost_ver_2(db):
    pass

#----------------------------------------------------------------------------------------
def upload_xtrachef_item_library(item_library_file):

    """ Take the downloaded item library from Xtrachef and modify it for uploading to database"""
    """ Make sure to save the downloaded .csv file into an .xlsx file, and format the columns."""

    # Create a Normalized Item Description column prior to manipulating data

    wb = load_workbook(item_library_file)
    ws = wb.active

        # Insert column C
    ws.insert_cols(5)

    # Set header
    ws.cell(row=1, column=5, value="Normalized Item Description")



    wb.save(f"✅ Normalized Item Description Column added to {item_library_file}")

    # Read the item_library_file and normalize the item_description columns and input that new data into that column in the database

    item_library_data = pd.read_excel(item_library_file)

    #print(item_library_data.columns.tolist())

    # Uses vectorized operations with pandas .apply() method to fill the empty normalized column
    item_library_data['Normalized Item Description'] = item_library_data['Item Description'].apply(normalize)
    # Uses vectorized operations with pandas .apply() method to normalize the vendor name
    item_library_data['Vendor Name'] = item_library_data['Vendor Name'].apply(normalize)
    # Make sure the "Last Purchased Date" column is a proper datetime object:
    item_library_data['Last Purchased Date'] = pd.to_datetime(item_library_data['Last Purchased Date'], errors='coerce')

    # This ensures that the most recent entry comes first.
    item_library_data = item_library_data.sort_values(by='Last Purchased Date', ascending=False)

    # This keeps only the first (i.e., most recent) row for each normalized description.
    deduped_data = item_library_data.drop_duplicates(subset='Normalized Item Description', keep='first')


    deduped_data.to_excel(item_library_file, index=False)
    print(f"✅ {item_library_file} has been normalized!")
