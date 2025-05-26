#Fills out a requisition form for AM prep team for events prep.
# Saved into the event_name folder as a separate xlsx file to be later uploaded to a shared google drive
#*** Need to add a boolean column in db for am prep
#*** Need to add sql query to check boolean column

import pandas as pd
from openpyxl import load_workbook #imports python library for reading and writting excel files
import sqlite3
from datetime import date
import os
import shutil
from excel_format import format_prep_sheet


def req_prep(item_ids, excel_folder_path, event_date, event_name, db):
    
    # Get the current working directory
    cwd = os.getcwd()
    file_count = 0
    # Create name for new event_req xlsx file
    new_file_name = f"PREP REQ_{event_name}_{event_date}_{file_count}.xlsx"

    # EVENTS REQ TEMPLATE file_path
    event_req_template_file_path = os.path.join(cwd, 'PREP REQ - TEMPLATE.xlsx')

    # File path for the folder where the copied template will be saved to
    dest_dir = os.path.join(cwd, excel_folder_path)

    # Complet file path of newly copied and renamed events req file > event folder
    dest_path = os.path.join(dest_dir, new_file_name)

    # Sanity check that the source exists
    if not os.path.isfile(event_req_template_file_path ):
        raise FileNotFoundError(f"Source file not found: {event_req_template_file_path }")
    
    while os.path.exists(dest_path):
        file_count += 1

        new_file_name = f"PREP REQ_{event_name}_{event_date}_{file_count}.xlsx"
        dest_path = os.path.join(dest_dir, new_file_name)

    # 5. Copy (using copy2 to preserve metadata)
    shutil.copy2(event_req_template_file_path , dest_path)

    print(f"Copied:\n  {event_req_template_file_path }\n→ {dest_path}")

    # Query db for all prep that can be requisitioned from AM prep team for Events
    conn = sqlite3.connect(db)
    # Cursor to execute commands
    cursor = conn.cursor()
    current_date = date.today()
    formatted_date = current_date.strftime("%m-%d-%Y")

    am_prep_req_list = []
    sous_prep_req_list = []
    for id in item_ids:
        cursor.execute(
                       f"""
                       SELECT req_prep.prep
                       FROM req_prep
                       JOIN menu_req_prep_list ON req_prep.req_prep_id = menu_req_prep_list.req_prep_id
                       WHERE menu_req_prep_list.menu_item_id = {id} AND req_prep.am_prep_team = 1;
                       """)  
        
        #.fetchall() is a list of tuples
        mise = cursor.fetchall()
        # access the tuple inside the list
        for tuple_item in mise:
             print(f"tuple_item{tuple_item}")
             am_prep_req_list.append(tuple_item[0])

    print( am_prep_req_list)

    # Populate new template with prep items that can be requisitioned from the AM Prep Team
    wb = load_workbook(f"{dest_dir}/{new_file_name}")
    ws = wb['AM Prep']
    ws['A1'] = f"AM EVENT PREP {formatted_date}"         
    print(am_prep_req_list)
    
    # Write each item into its own row (column A)
   # start = 3, becuase I want to start filling in the cells in the third row (rows 1-2 are titles and headings)
    for row_idx, prep_items in enumerate(am_prep_req_list, start=3):   # start=1 → Excel’s first row
        ws.cell(row=row_idx, column=1, value=prep_items)

    # format AM prep reauisition sheet
    format_prep_sheet (ws, 3, 1, 5)
    # Save
    wb.save(f"{dest_dir}/{new_file_name}")

    #Query db for all prep that only Sous Team / AM,PM line cookcs can do for Events

    for id in item_ids:
        cursor.execute(
                       f"""
                       SELECT req_prep.prep
                       FROM req_prep
                       JOIN menu_req_prep_list ON req_prep.req_prep_id = menu_req_prep_list.req_prep_id
                       WHERE menu_req_prep_list.menu_item_id = {id} AND req_prep.sous_prep = 1;
                       """)  
        
        #.fetchall() is a list of tuples
        mise = cursor.fetchall()
        # access the tuple inside the list
        for tuple_item in mise:
            sous_prep_req_list.append(tuple_item[0])

    wb = load_workbook(f"{dest_dir}/{new_file_name}")
    ws = wb['Sous Prep']
    ws['A1'] = f"SOUS EVENT PREP {formatted_date}"         
    print(sous_prep_req_list)

    # Write each item into its own row (column A)
    row_idx = 3
    for row_idx, prep_items in enumerate(sous_prep_req_list, start=3):   # start=1 → Excel’s first row
        ws.cell(row=row_idx, column=1, value=prep_items)

    # format AM prep reauisition sheet
    format_prep_sheet (ws, 3, 1, 5)

    # Save
    wb.save(f"{dest_dir}/{new_file_name}")

    
          
# ------------------------------------------------------------------------------------------
def test_prep_req(item_ids, db):

 # Query db for all prep that can be requisitioned from AM prep team for Events
    #unpacked_item_ids = [id for id in item_ids[0]]
    conn = sqlite3.connect(db)
    # Cursor to execute commands
    cursor = conn.cursor()

    am_prep_req_list = []
    sous_prep_req_list = []
    for id in item_ids:
        cursor.execute(
                       f"""
                       SELECT req_prep.prep
                       FROM req_prep
                       JOIN menu_req_prep_list ON req_prep.req_prep_id = menu_req_prep_list.req_prep_id
                       WHERE menu_req_prep_list.menu_item_id = {id} AND req_prep.am_prep_team = 1;
                       """)  
        
        #.fetchall() is a list of tuples
        mise = cursor.fetchall()
        # access the tuple inside the list
        for tuple_item in mise:
             print(f"tuple_item{tuple_item}")
             am_prep_req_list.append(tuple_item[0])

    print( f"am_prep_req_list: {am_prep_req_list}")

    for id in item_ids:
        cursor.execute(
                       f"""
                       SELECT req_prep.prep
                       FROM req_prep
                       JOIN menu_req_prep_list ON req_prep.req_prep_id = menu_req_prep_list.req_prep_id
                       WHERE menu_req_prep_list.menu_item_id = {id} AND req_prep.sous_prep = 1;
                       """)  
        
        #.fetchall() is a list of tuples
        mise = cursor.fetchall()
        # access the tuple inside the list
        for tuple_item in mise:
            sous_prep_req_list.append(tuple_item[0])

    print( f"sous_prep_req_list: {sous_prep_req_list}")
