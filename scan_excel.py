import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook





def update_product_catalog():
    # === File Paths ===
    source_file = "new_product_catalog_2025.xlsx"
    current_file = "current_product_catalog_2025.xlsx"
    template_file = "product_catalog_db_template.xlsx"
    output_file = "updated_product_catalog.xlsx"

    # === Step 1: Read Excel Files ===
    df_source = pd.read_excel(source_file, usecols=[0,1,2,5,11,13])  # Extract 6 columns
    df_current = pd.read_excel(current_file, usecols=[0,1,2,3,4,5,6,7])

    """Access a item_code columns from both df"""
    # source_item_code = df_source["Item Code"]
    # current_item_code = df_current["ingredient_code"]
    # current_ingredient_id = df_current["ingredient_id"]
    current_data = df_current[["ingredient_id", "purveyor", "ingredient_code", "ingredient_description",
                               "ingredient_name", "pack_size_unit", "purchase_price", "ingredient_type"]]
    source_data = df_source[["Item Description", "Vendor Name", "Item Code",
                               "Pack/Size/Unit", "Last Purchased Price ($)", "Product(s)"]]
    new_data_set =[]
    source_data_set = []

    for _, row in current_data.iterrows():
        new_data_set.append({
            "ingredient_id": row["ingredient_id"],
            "purveyor": row["purveyor"],
            "ingredient_code": row["ingredient_code"],
            "ingredient_description": row["ingredient_description"],
            "ingredient_name": row["ingredient_name"],
            "pack_size_unit": row["pack_size_unit"],
            "purchase_price": row["purchase_price"],
            "ingredient_type": row["ingredient_type"]
        })

    for _, row in source_data.iterrows():
        source_data_set.append({
            "ingredient_id": row["Item Description"],
            "purveyor": row["Vendor Name"],
            "ingredient_code": row["Item Code"],
            "ingredient_description": row["Item Description"],
            "pack_size_unit": row["Pack/Size/Unit"],
            "purchase_price": row["Last Purchased Price ($)"],
            "ingredient_type": row["Product(s)"]
        })

    # for new_dict_item in new_data_set:
    #     for source_dict_item in source_data_set:
    #         if new_dict_item["ingredient_code"] == source_dict_item["ingredient_code"]:
    #             new_dict_item["purveyor"] = source_dict_item["purveyor"]
    #             new_dict_item["ingredient_description"] = source_dict_item["ingredient_description"]
    #             new_dict_item["pack_size_unit"]  = source_dict_item["pack_size_unit"]
    #             new_dict_item["purchase_price"] = source_dict_item["purchase_price"]
    #             new_dict_item["ingredient_type"] = source_dict_item["ingredient_type"]


      # Step 1: Create source lookup dictionary
    source_lookup = {
        item["ingredient_code"]: item
        for item in source_data_set
    }

    # Step 2: Iterate through new data and update if matched
    for new_dict_item in new_data_set:
        code = new_dict_item["ingredient_code"]
        if code in source_lookup:
            match = source_lookup[code]
            new_dict_item.update({
                "purveyor": match["purveyor"],
                "ingredient_description": match["ingredient_description"],
                "pack_size_unit": match["pack_size_unit"],
                "purchase_price": match["purchase_price"],
                "ingredient_type": match["ingredient_type"]
            })

    print(new_data_set)

    wb = load_workbook(output_file)
    ws = wb['ingredients']
#    # Write each item into its own row 
    for row_idx, dict_items in enumerate(new_data_set, start=2):   # start=1 → Excel’s first row
        ws.cell(row=row_idx, column=1, value=dict_items["ingredient_id"])
        ws.cell(row=row_idx, column=2, value=str(dict_items["purveyor"]))
        ws.cell(row=row_idx, column=3, value=dict_items["ingredient_code"])
        ws.cell(row=row_idx, column=4, value=dict_items["ingredient_description"])
        ws.cell(row=row_idx, column=5, value=dict_items["ingredient_name"])
        ws.cell(row=row_idx, column=6, value=dict_items["pack_size_unit"])
        ws.cell(row=row_idx, column=7, value=dict_items["purchase_price"])
        ws.cell(row=row_idx, column=8, value=dict_items["ingredient_type"])

 
        
    
    wb.save(output_file)

    print(f"✅ Template filled and saved as: {output_file}")
    

    # # Define fuzzy matching function
    # def fuzzy_match(value, comparison_series, scorer=fuzz.token_sort_ratio, threshold=80):
    #     if pd.isna(value):
    #         return None
    #     # Compare against all current ingredients
    #     result = process.extractOne(str(value), comparison_series.dropna().astype(str).tolist(), scorer=scorer)
    #     if result:
    #         match, score = result[:2]
    #         return match if score >= threshold else None
    #     return None
    


 


    # Apply fuzzy match to each row in source_ingredient
    # updated_ingredient =  current_item_code.apply(
    # lambda x: fuzzy_match(x,source_item_code)
    # )
    # #print(updated_ingredient)
  

    # # === Step 5: Load Template and Insert Data ===
    # wb = load_workbook(template_file)
    # ws = wb.active  # You can specify sheet name with wb['SheetName']

    # #Start writing from row 2 (assuming row 1 is header)
    # start_row = 2

    # for idx, value in updated_ingredient.items():
    #     ws.cell(row=start_row + idx, column=3, value=value)  # Write into first column

    # df_new = pd.read_excel(output_file, usecols=[0,1,2,3,4,5,6,7])

    # updated_ingredient_item_code = []
    # for idx, row in df_new.iterrows():
    #     for idx, row in df_source:
        
            

 

    # # === Step 6: Save the filled template ===
    # wb.save(output_file)

    #print(f"✅ Template filled and saved as: {output_file}")