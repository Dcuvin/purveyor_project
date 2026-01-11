
from openpyxl import load_workbook #imports python library for reading and writting excel files
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def format_headers_and_borders(sheet, start_row, start_col, end_col):
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        
    # Define the font for non-header cells
    cell_font = Font(name="Calibri", size=14)

    # Define the alignment for all cells
    cell_alignment = Alignment(horizontal='center', vertical='center')

        
    # Apply font to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = cell_font
            # Center all data
            cell.alignment = cell_alignment
    # Format headers
    for cell in sheet.iter_cols(min_row=start_row, max_row=start_row, min_col=start_col, max_col=end_col):
        for c in cell:
            c.font = Font(bold=True, name='Calibri', size=14, color="000000")
            c.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
            c.border = thin_border

    # Apply borders to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):      
        for cell in row:
            cell.border = thin_border
               
                
# Function to set print options
def set_print_options(sheet):
        sheet.print_options.gridLines = False
        sheet.page_setup.orientation = 'portrait'
        # Set print settings to fit on one page
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1  # 0 means "as many as needed"
    
# Function to insert unformatted rows
def insert_blank_rows(sheet, start_row):
    sheet.insert_rows(start_row, 1)

# Function that formats the order sheet

def format_order_sheet(sheet, start_row, start_col, end_col):
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        
    # Define the font for non-header cells
    cell_font = Font(name="Calibri", size=14)

    # Define the alignment for all cells
    cell_alignment = Alignment(horizontal='center', vertical='center')

        
    # Apply font to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = cell_font
            # Center all data
            cell.alignment = cell_alignment
    # Format headers for cell 'A1' / top left cell only.
    for cell in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=1):
        for c in cell:
            c.font = Font(bold=True, name='Calibri', size=16, color="000000")
            c.border = thin_border
    # Format headers for row 3 / the first row only.
    for cell in sheet.iter_cols(min_row=3, max_row=3, min_col=1, max_col=3):
        for c in cell:
            c.font = Font(bold=True, name='Calibri', size=14, color="000000")
            c.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
            c.border = thin_border

    # Apply borders to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):      
        for cell in row:
            cell.border = thin_border
               
#----------------------------------------------------------------------------

def format_prep_sheet (sheet, start_row, start_col, end_col):

    # Define the font for non-header cells
    cell_font = Font(name="Calibri", size=14)

    # Define the alignment for all cells
    cell_alignment = Alignment(horizontal='center', vertical='center')

        
    # Apply font to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = cell_font
            # Center all data
            cell.alignment = cell_alignment
#----------------------------------------------------------------------------

def format_table(ws, start_row, start_col, dataframe):
    num_rows = dataframe.shape[0]
    num_cols = dataframe.shape[1]

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Use black font for both header and body
    header_font = Font(name='Calibri', size=14,bold=True, color="FF000000")  # Black, bold
    body_font = Font(name='Calibri', size=12, color="FF000000")  # Black
    align_center = Alignment(horizontal='center', vertical='center')

    # Format header row
    # Need to hardcode range(2) as the 'Need'(as in how much mise do you need to make ex: 1 quart, 30 portions, etc.) column is left empty for later manual input.
    for col_offset in range(2):
        col_letter = get_column_letter(start_col + col_offset)  # openpyxl is 1-based
        cell = ws[f"{col_letter}{start_row + 1}"]  # pandas writes header here
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
        #print("Formatting column:", col_letter)


    # Format body rows
    for row_offset in range(num_rows):
        for col_offset in range(2):
            col_letter = get_column_letter(start_col + col_offset)
            cell = ws[f"{col_letter}{start_row + 2 + row_offset}"]  # data starts after header
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = align_center

#----------------------------------------------------------------------------

def format_table_ver_2(ws, start_row, start_col, dataframe):
    # num_rows = dataframe.shape[0]
    # num_cols = dataframe.shape[1]
    df_length =len(dataframe)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Use black font for both header and body
    header_font = Font(name='Calibri', size=14,bold=True, color="FF000000")  # Black, bold
    body_font = Font(name='Calibri', size=14, color="FF000000")  # Black
    align_center = Alignment(horizontal='center', vertical='center')

    # Format header row
    for col in range(6):
        col_letter = get_column_letter(col + 1)  # openpyxl is 1-based
        cell = ws[f"{col_letter}{start_row + 1}"]  # pandas writes header here
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = align_center
        cell.fill = PatternFill(start_color="FFC9DAF8", end_color="FFC9DAF8", fill_type="solid")
     
        #print("Formatting column:", col_letter)


    # # Format body rows
    for row in range(df_length):
        for col in range(6):
            col_letter = get_column_letter(col + 1)
            cell = ws[f"{col_letter}{start_row + 2 + row}"]  # data starts after header
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = align_center
#----------------------------------------------------------------------------
def format_order_guide (sheet, start_row, start_col, end_col):

    # Define the font for non-header cells
    cell_font = Font(name="Calibri", size=14)

    # Define the alignment for all cells
    cell_alignment = Alignment(horizontal='center', vertical='center')

        
    # Apply font to the entire table
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.font = cell_font
            # Center all data
            cell.alignment = cell_alignment